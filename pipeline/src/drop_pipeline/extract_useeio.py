"""Extract USEEIO v2.0 sector freshwater-withdrawal intensities.

Sheet N = direct + indirect impact coefficients per 2012 USD producer price.
We convert to litres per purchaser dollar in the latest price year available
(Phi = producer share of purchaser price; Rho = producer price-year ratio to
2012, so year-Y dollars / rho = 2012 dollars).

USEEIO v2.5.1 has NO water indicator; this script refuses to read it and
asserts that stays true so nobody wires it in by mistake.
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import yaml

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from drop_pipeline.emit import DATASETS, write_json

V20 = DATASETS / "useeio" / "USEEIOv2.0.xlsx"
V251 = DATASETS / "useeio" / "USEEIOv2.5.1-waxwing-22.xlsx"
CONFIG = Path(__file__).resolve().parents[2] / "config"

WATER_ROW = "Freshwater withdrawals"


def refuse_v251():
    wb = openpyxl.load_workbook(V251, read_only=True)
    names = set(wb.sheetnames)
    has_watr = False
    if "indicators" in names:
        for row in wb["indicators"].iter_rows(values_only=True):
            if any(c and "WATR" in str(c) for c in row):
                has_watr = True
                break
    wb.close()
    if has_watr:
        raise SystemExit(
            "USEEIOv2.5.1 unexpectedly contains a WATR indicator — "
            "re-evaluate which model release feeds the water table.")
    print("v2.5.1 has no WATR indicator — refused as a water source (correct)")


def sheet_matrix_row(wb, sheet, row_label):
    ws = wb[sheet]
    it = ws.iter_rows(values_only=True)
    header = next(it)
    cols = [str(c) for c in header[1:] if c is not None]
    for row in it:
        if row[0] and str(row[0]).strip() == row_label:
            return cols, {c: float(v) for c, v in zip(cols, row[1:])
                          if v is not None}
    raise SystemExit(f"row {row_label!r} not found in sheet {sheet}")


def year_matrix(wb, sheet):
    """Rho/Phi: rows = commodity IDs, cols = years."""
    ws = wb[sheet]
    it = ws.iter_rows(values_only=True)
    years = [str(c) for c in next(it)[1:] if c is not None]
    out = {}
    for row in it:
        if row[0] is None:
            break
        out[str(row[0])] = {y: float(v) for y, v in zip(years, row[1:])
                            if v is not None}
    return years, out


def main():
    refuse_v251()
    consumer = yaml.safe_load(
        (CONFIG / "useeio_consumer_sectors.yaml").read_text())["sectors"]
    consumer = {str(k): v for k, v in consumer.items()}

    wb = openpyxl.load_workbook(V20, read_only=True)

    # indicator sanity
    ok = any(
        str(r[3]) == "WATR" and str(r[2]) == WATER_ROW
        for r in wb["indicators"].iter_rows(min_row=2, values_only=True)
        if r[0] is not None
    )
    if not ok:
        raise SystemExit("WATR indicator not found in v2.0")

    cols, water = sheet_matrix_row(wb, "N", WATER_ROW)
    meta = {}
    for r in wb["commodities_meta"].iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            break
        meta[str(r[1])] = {"name": str(r[2]), "code": str(r[3]),
                           "location": str(r[4]),
                           "description": str(r[5]) if r[5] else None}
    rho_years, rho = year_matrix(wb, "Rho")
    phi_years, phi = year_matrix(wb, "Phi")
    latest = max(set(rho_years) & set(phi_years), key=int)

    sectors = []
    for cid in cols:
        m = meta.get(cid)
        if m is None or cid not in water:
            continue
        code = m["code"]
        n_kg_per_usd2012 = water[cid]
        phi_l = phi.get(cid, {}).get(latest)
        rho_l = rho.get(cid, {}).get(latest)
        purchaser = (n_kg_per_usd2012 * phi_l / rho_l
                     if phi_l and rho_l else None)
        entry = {
            "factor_id": f"useeio:{cid}",
            "dataset": "useeio_v2.0",
            "dataset_release": "USEEIO v2.0 (US EPA)",
            "sector_code": code,
            "sector_id": cid,
            "name": m["name"],
            "location": m["location"],
            "metric_type": "freshwater_withdrawal",
            "proxy_metric": True,
            "value_l_per_usd_2012_producer": n_kg_per_usd2012,
            "value_l_per_usd_purchaser": purchaser,
            "price_year": int(latest),
            "phi": phi_l,
            "rho": rho_l,
            "geography": "US",
            "system_boundary":
                "cradle-to-producer-gate, US economy with import proxy",
            "confidence_cap": "very_low",
            "rights": "US EPA, public domain",
        }
        if code in consumer:
            entry["consumer"] = {
                "display": consumer[code]["display"],
                "typical_price_usd": consumer[code]["typical_price_usd"],
                "synonyms": consumer[code].get("synonyms", []),
            }
        sectors.append(entry)

    write_json("sector_useeio.json", {"useeio_sectors": sectors})

    by_code = {s["sector_code"]: s for s in sectors}
    assert len(sectors) == 411, len(sectors)
    assert abs(by_code["315000"]["value_l_per_usd_2012_producer"] - 20.813) < 0.01
    assert abs(by_code["334111"]["value_l_per_usd_2012_producer"] - 9.181) < 0.01
    assert abs(by_code["322291"]["value_l_per_usd_2012_producer"] - 25.243) < 0.01
    marked = [s for s in sectors if "consumer" in s]
    missing = set(consumer) - {s["sector_code"] for s in marked}
    assert not missing, f"consumer sectors not found: {missing}"
    print(f"useeio: {len(sectors)} sectors ({len(marked)} consumer-curated), "
          f"price year {latest} — gates passed")


if __name__ == "__main__":
    main()
